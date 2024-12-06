import asyncio
import aiohttp
import os
import json
from lxml import html
import openpyxl
import traceback
from time import sleep
import random
from playwright.async_api import async_playwright
from PIL import Image, ImageDraw, ImageFont

mang = [0.5,0.9,1.3,1.9]
mang2 = [0.5,0.9,1.3]

async def download_image1(session, code, url, save_dir, semaphore):
    async with semaphore:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        }
        try:
            async with session.get(url,  headers=headers) as response:
                
                    if response.status == 200 and r'#colors-' in url:
                        sleep(random.choice(mang))
                        newUrl = url.split(r'#colors-')[0]
                        number = url.split(r'#colors-')[1]
                        print(f"new: {newUrl}, {number}")
                        async with session.get(newUrl,  headers=headers) as new_response:
                            if new_response.status == 200:
                                html_content = await new_response.text()
                                tree = html.fromstring(html_content)
                                imgs = tree.xpath("//div[@class='col-image-15']/a")
                                print(number)
                                print(newUrl)
                                data_img = imgs[int(number)-1].get('href')
                                
                                # Tải và lưu hình ảnh
                                async with session.get(data_img,  headers=headers) as img_response:
                                    if img_response.status == 200:
                                        image_content = await img_response.read()  # Đọc nội dung hình ảnh
                                        # Lưu hình ảnh vào tệp
                                        with open(f"{save_dir}{code}.jpg", "wb") as file:
                                             file.write(image_content)
                                        print(f"Hình ảnh đã được lưu tại: D:\\IT-Only\\python\\playw2\\anh\\midamericacomponents\\{code}.jpg")
                                        return True
                                    else:
                                        print(f"Lỗi khi tải hình ảnh: {img_response.status}")
                            else:
                                        print(f"Lỗi khi tải Link ảnh: {img_response.status}")
                    else:
                        sleep(random.choice(mang2))
                        print(f"Failed to download {url}: {response.status}")

        except Exception  as e:
            print(f"Loi Khac: {e}")
            return False

async def download_image2(session,codex, url,save_dir, semaphore):
    async with semaphore:  # Giới hạn số lượng tác vụ chạy đồng thời
        try:
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=False, slow_mo=100)
                page = await browser.new_page()
                await page.goto(url, timeout=30000)
                # Cuộn xuống cuối trang bằng cách sử dụng mouse
                await page.mouse.wheel(0, 1000)  # Cuộn xuống
                await page.wait_for_timeout(3000)  # Đợi một chút để trang tải thêm nội dung nếu có
                # Processing
                html_content = await page.content()
                #print(await page.content())
                tree = html.fromstring(html_content)
                imgs = tree.xpath("//div[starts-with(@class, 'sc-dmyCSP')]")
                data_img = ''.join(imgs[0].get('style').split(':')[1]).strip().split(';')[0].split('(')
                print(data_img[0])
                print(data_img[1].split(')')[0])
                code = data_img[0].upper()
                color = data_img[1].split(')')[0].split(',')
                # Màu nền từ mã RGB
                rgb_color = (int(color[0]), int(color[1]), int(color[2]))

                # Kích thước hình ảnh
                width, height = 800, 800

                # Tạo hình ảnh mới với màu nền
                image = Image.new(code, (width, height), rgb_color)

                # Tạo đối tượng vẽ
                draw = ImageDraw.Draw(image)

                # Thêm văn bản
                text_color = (0, 0, 0)  # Màu chữ đen
                font = ImageFont.load_default()  # Sử dụng font mặc định

                # Lưu hình ảnh
                image.save(f"{save_dir}{codex}.jpg")
                await browser.close()
                return True
        except Exception  as e:
            print(f"Loi Khac: {e}")
            return False
async def download_image3(session, code, url, save_dir, semaphore):
    async with semaphore:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36" 
        }
        try:
            async with session.get(url,  headers=headers) as response:
                    if response.status == 200:
                        sleep(random.choice(mang))
                        html_content = await response.text()
                        #print(f"Status: 200 \n {html_content}")
                        sizes = r'/?wid=1024&fmt=webp-alpha&qlt=90&fit=hfit,1'
                        tree = html.fromstring(html_content)
                        imgs = tree.xpath("//div[starts-with(@class, 'image__base image1715272685269')]")
                        data_img = imgs[0].get('data-components-params-image')
                        print(data_img)

                        para_img = json.loads(data_img)
                        src_img = para_img.get('src')
                        full = src_img+sizes
                        print(full)
                        
                        # Tải và lưu hình ảnh
                        async with session.get(full,  headers=headers) as img_response:
                            if img_response.status == 200:
                                image_content = await img_response.read()  # Đọc nội dung hình ảnh
                                # Lưu hình ảnh vào tệp
                                with open(f"{save_dir}{code}.jpg", "wb") as file:
                                    file.write(image_content)
                                
                                print(f"Hình ảnh đã được lưu tại: D:\IT-Only\python\playw2\anh\OSI\{code}.jpg")
                                return True
                            else:
                                print(f"Lỗi khi tải hình ảnh: {img_response.status}")
                    else:
                            sleep(random.choice(mang2))
                            print(f"Failed to download {url}: {response.status}")
        except Exception  as e:
            print(f"Loi Khac: {e}")
            return False
            
async def download_imagex(session,code, url,save_dir, semaphore):
    async with session.get(url) as response:
            if response.status == 200:
                    html_content = await response.text()
                    #print(f"Status: 200 \n {html_content}")
                    sizes = r'/?wid=1024&fmt=webp-alpha&qlt=90&fit=hfit,1'
                    tree = html.fromstring(html_content)
                    skus = tree.xpath("//div[@id='text-a0db350e3a']")
                    sku = ' '.join(skus[0].text_content().split()).strip()
                    print(' '.join(skus[0].text_content().split()).strip())
                    imgs = tree.xpath("//div[starts-with(@class, 'image__base image1715272685269')]")
                    data_img = imgs[0].get('data-components-params-image')
                    print(data_img)

                    para_img = json.loads(data_img)
                    src_img = para_img.get('src')
                    full = src_img+sizes
                    print(full)
                    
                    # Tải và lưu hình ảnh
                    async with session.get(full) as img_response:
                        if img_response.status == 200:
                            image_content = await img_response.read()  # Đọc nội dung hình ảnh
                            # Lưu hình ảnh vào tệp
                            with open(f"D:\\IT-Only\\python\\playw2\\anh\\OSI\\{sku}.jpg", "wb") as file:
                                 file.write(image_content)
                            print(f"Hình ảnh đã được lưu tại: D:\\IT-Only\\python\\playw2\\anh\\OSI\\{sku}.jpg")
                        else:
                            print(f"Lỗi khi tải hình ảnh: {img_response.status}")
                    
            else:
                    print(f"Failed to download {url}: {response.status}")

def check(file_path,arrLink):
    # Kiểm tra xem tệp có tồn tại không
    if os.path.exists(file_path):
        print("Tệp Excel đã tồn tại. Đang mở tệp...")
        # Mở tệp Excel
        workbook = openpyxl.load_workbook(file_path)
        shetname  = []
        try:
            # Lặp qua từng sheet trong workbook
            for sheet_name in workbook.sheetnames:
                print(f"Đang kiểm tra sheet: {sheet_name}")
                shetname.append(sheet_name)
                sheet = workbook[sheet_name]
                # In số lượng hàng và cột
                print(f"Số lượng hàng: {sheet.max_row}, Số lượng cột: {sheet.max_column}")
                # Lặp qua các hàng và in nội dung
                arrLinks= {}
                for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_index == 0:
                        continue
                    arrLinks[f"{row[0]}"] = row[3]
                    #print(row)  # In ra từng hàng
                print(f"{arrLinks}")
                arrLink.append(arrLinks)

        finally:
            # Đảm bảo workbook được đóng
            workbook.close()
    else:
        print("Tệp Excel không tồn tại.")
    return shetname, arrLink
    
async def main():
    arrLink = []
    # Đường dẫn tới tệp Excel
    #file_path = r'D:\IT-Only\python\playw2\3webs.xlsx'  # Thay thế bằng đường dẫn của bạn
    file_path = input("Nhập đường dẫn file Excel: ")
    shetname , arrLinks = check(file_path,arrLink)
    print(arrLinks[1])

    semaphore = asyncio.Semaphore(5)
    current_path = os.getcwd()
    save_dir = f"{current_path}\\anh"
    
    os.makedirs(save_dir, exist_ok=True)
    save_dir = save_dir.replace("\\", "\\\\")
    
    completed = set()

    # Đọc tệp trạng thái nếu có
    if os.path.exists('completed_links.json'):
        with open('completed_links.json', 'r') as f:
            completed = set(json.load(f))
    async with aiohttp.ClientSession() as session:
            tasks = []
            urls_to_process = []
            
            cacheClear = input("Xóa cache không? Y/N: ")
            if cacheClear == 'y' or cacheClear == 'Y':
                data = json.loads('["one"]')                
                with open('completed_links.json', 'w') as json_file:
                    json.dump(data, json_file, indent=4)
            else:
                print("Not clear cache!")                
            
            n = int(input("Nhập hàm cần chạy 1/2/3: "))
            
            if n==1:
                funcNeed = download_image1
                indexUrl = 0
                save_dir = save_dir+ "\\\\" + shetname[0] + "\\\\"
                save_dirx = save_dir.replace("\\\\", "\\")
                os.makedirs(save_dirx, exist_ok=True)
            elif n==2:
                funcNeed = download_image2
                indexUrl = 1
                save_dir =save_dir+ "\\\\" + shetname[1] + "\\\\"                
                save_dirx = save_dir.replace("\\\\", "\\")
                os.makedirs(save_dirx, exist_ok=True)
            elif n==3:
                funcNeed = download_image3
                indexUrl = 2
                save_dir = save_dir+ "\\\\" + shetname[2] + "\\\\"
                
                save_dirx = save_dir.replace("\\\\", "\\")
                os.makedirs(save_dirx, exist_ok=True)
            n=input(save_dir)
            for code , url in arrLinks[indexUrl].items():
                if url not in completed:
                    tasks.append(funcNeed(session,code, url,save_dir, semaphore))
                    urls_to_process.append(url)
            results = await asyncio.gather(*tasks)
            # Cập nhật các yêu cầu đã hoàn thành
            for url, success in zip(urls_to_process, results):
                if success:
                    completed.add(url)
    # Ghi lại trạng thái vào tệp
    with open('completed_links.json', 'w') as f:
        json.dump(list(completed), f)
        print("Ghi xong!!!")
if __name__ == "__main__":
    asyncio.run(main())
